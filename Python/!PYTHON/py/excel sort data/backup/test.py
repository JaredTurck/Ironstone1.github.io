import openpyxl

sheet_object = openpyxl.load_workbook("test.xlsx").active

start_row = 2
float(start_row)
cell_object = sheet_object.cell(row = start_row, column = 2)
print(cell_object.value)

while cell_object.value is not None:
    start_row = start_row + 1
    cell_object = sheet_object.cell(row = start_row, column = 2)
    print(cell_object.value)
